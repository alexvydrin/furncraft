import openpyxl

# import os

"""
функции для работы с excel-файлом паспорта
"""

# TODO: вынести функцию в общие - она используется в нескольких модулях
"""
def del_double_space(s):
    while "  " in s:
        s = s.replace("  ", " ")
    return s
"""


def get_amount_from_passport(name_file, log):
    wb = openpyxl.load_workbook(filename=f"media/{name_file}", read_only=False, data_only=True)

    wb.active = 0  # выбираем первый лист
    ws = wb.active

    rows = ws.max_row
    columns = ws.max_column
    if rows == 1 and columns == 1:
        log.append(f"На первом листе в паспорте ничего нет")
        return []
    if ws.title.lower() == "сводный":
        log.append(f"Для сводного паспорта тестирование не реализовано")
        return []
    if ws.title.lower() == "фурнитура":
        log.append(f"Для отдельного списка фурнитуры тестирование не реализовано")
        return []
    materials = get_count_from_passport(ws, log)
    if not len(materials):
        log.append(f"Не могу узнать количество материалов в паспорте на листе: {ws.title}")

    return materials


def get_count_from_passport(ws, log):
    """
        :param ws: лист на котором размещен паспорт
        :param log: лог ошибок
        :return: список словарей с полями name и amount
        функция возвращает количество материалов в паспорте
        если находится ошибка, то информация о ней записывается в лог
    """

    # Ищем метку: ЛДСП
    r_ldsp, c_ldsp = find_mark("ЛДСП", ws, 5, 15, 2, 4, only_at_first=True)
    if not r_ldsp:
        log.append(f"Не могу найти метку 'ЛДСП': {ws.title}")
        return []

    # Ищем метку: Габариты ЛДСП
    cell_a_value = ws.cell(row=r_ldsp + 1, column=c_ldsp + 1).value
    cell_b_value = ws.cell(row=r_ldsp + 1, column=c_ldsp + 2).value
    if not (isinstance(cell_a_value, str) and isinstance(cell_b_value, str)):
        log.append(f"Не могу найти метку 'Габариты ЛДСП (ошибка типа)': {ws.title}")
        return []
    if not ("Габариты" in ws.cell(row=r_ldsp, column=c_ldsp + 1).value and
            "A" in cell_a_value and
            "B" in cell_b_value):
        log.append(f"Не могу найти метку 'Габариты ЛДСП': {ws.title}")
        return []
    c_ldsp_a = c_ldsp + 1
    c_ldsp_b = c_ldsp + 2

    # Ищем метку: Количество ЛДСП
    if "Кол" not in ws.cell(row=r_ldsp, column=c_ldsp + 3).value:
        log.append(f"Не могу найти метку 'Количество ЛДСП': {ws.title}")
        return []
    c_ldsp_count = c_ldsp + 3

    # Ищем схему кромления 1 типа (тонкая кромка)
    r_edge_1, c_edge_1 = find_mark("Кромка", ws, r_ldsp, r_ldsp, c_ldsp_count + 1, c_ldsp_count + 2)
    if not r_edge_1:
        log.append(f"Не могу найти схему кромления 1 типа: {ws.title}")
        return []
    if not ("A" in ws.cell(row=r_ldsp + 1, column=c_edge_1 + 0).value and
            "B" in ws.cell(row=r_ldsp + 1, column=c_edge_1 + 2).value):
        log.append(f"Не могу найти A и B в схеме кромления 1 типа: {ws.title}")
        return []
    c_edge_1_a = c_edge_1 + 0
    c_edge_1_b = c_edge_1 + 2

    right_coord = c_edge_1_b

    # Ищем схему кромления 2 типа (толстая кромка) - может и не быть
    r_edge_2, c_edge_2 = find_mark("ПВХ", ws, r_ldsp, r_ldsp, c_edge_1_b + 1, c_edge_1_b + 1)
    if not r_edge_2:
        # log.append(f"Не могу найти схему кромления 2 типа: {f} {ws.title}")
        c_edge_2_a = False
        c_edge_2_b = False
    else:
        if not ("A" in ws.cell(row=r_ldsp + 1, column=c_edge_2 + 0).value and
                "B" in ws.cell(row=r_ldsp + 1, column=c_edge_2 + 2).value):
            log.append(f"Не могу найти A и B в схеме кромления 2 типа: {ws.title}")
            return []
        c_edge_2_a = c_edge_2 + 0
        c_edge_2_b = c_edge_2 + 2
        right_coord = c_edge_2_b

    # Ищем результаты расчетов
    # Сначала ищем расход ЛДСП
    r_ldsp_size, c_ldsp_size = find_mark("Расход", ws, r_ldsp, r_ldsp, right_coord + 1, right_coord + 1)
    if not r_ldsp_size:
        log.append(f"Не могу найти размеры ЛДСП: {ws.title}")
        return []
    # ищем расход кромки 1
    r_edge_1_size, c_edge_1_size = find_mark("Кромка", ws, r_ldsp, r_ldsp, c_ldsp_size + 1, c_ldsp_size + 1)
    if not r_edge_1_size:
        log.append(f"Не могу найти расход кромки 1: {ws.title}")
        return []
    # ищем расход кромки 2
    r_edge_2_size, c_edge_2_size = False, False
    if r_edge_2:
        r_edge_2_size, c_edge_2_size = find_mark("ПВХ", ws, r_ldsp, r_ldsp, c_edge_1_size + 1, c_edge_1_size + 1)
        if not r_edge_2_size:
            # log.append(f"Не могу найти расход кромки 2: {f} {ws.title}")
            # это нормальная ситуация, возможно ПВХ не используется - примем расход за 0
            # return []
            pass

    # Проверяем формулы в строках
    r = r_ldsp + 2  # Первая строка в таблице ЛДСП
    ldsp_count_total, ldsp_size_total, edge_1_size_total, edge_2_size_total = 0, 0, 0, 0
    while ws.cell(row=r, column=c_ldsp_a).value or ws.cell(row=r, column=c_ldsp - 1).value:
        ldsp_a = ws.cell(row=r, column=c_ldsp_a).value or 0
        ldsp_b = ws.cell(row=r, column=c_ldsp_b).value or 0
        ldsp_count = ws.cell(row=r, column=c_ldsp_count).value or 0
        edge_1_a = ws.cell(row=r, column=c_edge_1_a).value or 0
        edge_1_b = ws.cell(row=r, column=c_edge_1_b).value or 0
        if r_edge_2:
            edge_2_a = ws.cell(row=r, column=c_edge_2_a).value or 0
            edge_2_b = ws.cell(row=r, column=c_edge_2_b).value or 0
        else:
            edge_2_a = 0
            edge_2_b = 0
        ldsp_size = ws.cell(row=r, column=c_ldsp_size).value or 0
        edge_1_size = ws.cell(row=r, column=c_edge_1_size).value or 0
        if c_edge_2_size:
            edge_2_size = ws.cell(row=r, column=c_edge_2_size).value or 0
        else:  # возможно кромка 2 не используется
            edge_2_size = 0
        ldsp_count_total += ldsp_count

        ldsp_size_item = ldsp_a * ldsp_b / 1000000 * ldsp_count
        if ldsp_size_item != ldsp_size:
            log.append(f"Ошибка в расчете размера ЛДСП строка = {r} ({ldsp_size_item} != {ldsp_size}) {ws.title}")
        ldsp_size_total += ldsp_size_item

        edge_1_size_item = 0
        if edge_1_a or edge_1_b:
            edge_1_a = 0 if not edge_1_a else edge_1_a
            edge_1_b = 0 if not edge_1_b else edge_1_b
            edge_1_size_item = (ldsp_a * edge_1_a + ldsp_b * edge_1_b) / 1000 * ldsp_count
            if edge_1_size_item != edge_1_size:
                log.append(f"Ошибка в расчете кромки 1 строка = {r} ({edge_1_size_item} != {edge_1_size}) {ws.title}")
        edge_1_size_total += edge_1_size_item

        edge_2_size_item = 0
        if edge_2_a or edge_2_b:
            edge_2_a = 0 if not edge_2_a else edge_2_a
            edge_2_b = 0 if not edge_2_b else edge_2_b
            edge_2_size_item = (ldsp_a * edge_2_a + ldsp_b * edge_2_b) / 1000 * ldsp_count
            if edge_2_size_item != edge_2_size:
                log.append(f"Ошибка в расчете кромки 2 строка = {r} ({edge_2_size_item} != {edge_2_size})"
                           f" {ws.title}")
        edge_2_size_total += edge_2_size_item

        r += 1

    ldsp_count_total_fact = ws.cell(row=r, column=c_ldsp_count).value or 0
    ldsp_size_total_fact = ws.cell(row=r, column=c_ldsp_size).value or 0
    edge_1_size_total_fact = ws.cell(row=r, column=c_edge_1_size).value or 0
    if c_edge_2_size:
        edge_2_size_total_fact = ws.cell(row=r, column=c_edge_2_size).value or 0
    else:  # возможно кромка 2 не используется
        edge_2_size_total_fact = 0

    if ldsp_count_total != ldsp_count_total_fact:
        log.append(f"Ошибка в расчете итогового количества деталей ({ldsp_count_total} != {ldsp_count_total_fact})"
                   f" {ws.title}")
    if ldsp_size_total != ldsp_size_total_fact:
        log.append(f"Ошибка в расчете итогового размера ЛДСП ({ldsp_size_total} != {ldsp_size_total_fact})"
                   f" {ws.title}")
    if edge_1_size_total != edge_1_size_total_fact:
        log.append(f"Ошибка в расчете итогового расхода кромки 1 ({edge_1_size_total} != {edge_1_size_total_fact})"
                   f" {ws.title}")
    if edge_2_size_total != edge_2_size_total_fact:
        log.append(f"Ошибка в расчете итогового расхода кромки 2 ({edge_2_size_total} != {edge_2_size_total_fact})"
                   f" {ws.title}")

    ret = []
    material = {"name": "лдсп 16мм", "amount": ldsp_size_total_fact}
    ret.append(material)
    material = {"name": "пвх 19х0,4мм", "amount": edge_1_size_total_fact}
    ret.append(material)
    material = {"name": "пвх 19х2,0мм", "amount": edge_2_size_total_fact}
    ret.append(material)

    # Ищем информацию о прочей фурнитуре

    r, c = find_mark("фурнитура", ws, r + 1, ws.max_row - 1, c_ldsp, c_ldsp)
    if not r:
        log.append(f"Не могу найти метку 'фурнитура': {ws.title}")
        return []

    r += 1  # Первая строка в таблице фурнитуры
    while ws.cell(row=r, column=c).value:
        mat = ws.cell(row=r, column=c + 0).value
        cnt = ws.cell(row=r, column=c + 1).value or 0
        if mat and cnt:
            material = {"name": mat, "amount": cnt}
            ret.append(material)
        r += 1

    return ret


def find_mark(mark, ws, r_start, r_finish, c_start, c_finish, only_at_first=False):
    """
    :param mark: метка для поиска
    :param ws: лист на котором ищется метка
    :param r_start: начальная строка
    :param r_finish: конечная строка
    :param c_start: начальная колонка
    :param c_finish: конечная колонка
    :param only_at_first: True если искать только в начале строки
    :return: кортеж row, column - координаты найденной метки, false если метка не найдена
    """

    row = False
    column = False
    for r in range(r_start, r_finish + 1):
        for c in range(c_start, c_finish + 1):
            value = ws.cell(row=r, column=c).value
            if isinstance(value, str):
                l_ok = False
                if only_at_first:
                    if mark.lower() == value.lower()[:len(mark)]:
                        l_ok = True
                else:  # not only_at_first:
                    if mark.lower() in value.lower():
                        l_ok = True
                if l_ok:
                    row = r
                    column = c
                    break
        if row:
            break
    return row, column
