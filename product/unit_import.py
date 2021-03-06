import openpyxl

"""
функции для работы с excel-файлами
    - прайс - price_main.xlsx
    - расчет себестоимости - себест общ новая.xlsx
1) импорт данных из excel-файлов
2) тестирование идентичности данных из excel-файлов и в базе данных
данные:
    1) список изделий
    2) список статей затрат
    3) строки калькуляции - расход для изделия по статьям затрат (количество или сумма)    

+) функция пересчета расчетной цены

"""


def get_setting_rewrite():
    from product.models import Settings
    return Settings.objects.get(name="main").is_rewriting


def del_double_space(s):
    while "  " in s:
        s = s.replace("  ", " ")
    return s


def get_d_pricelist():

    # f = r"static/data/Прайс Основной 2019октябрь_магазины.xlsx"
    f = r"static/data/price_main.xlsx"

    d_pricelist = []

    wb = openpyxl.load_workbook(filename=f, read_only=False, data_only=True)

    ws = wb['прайс']
    rows = ws.max_row
    for c in (4, 5):  # колонки эко и лайт
        for r in range(5, rows + 1):
            name = ws.cell(row=r, column=3).value
            if not name or not isinstance(name, str):
                continue
            name = del_double_space(name.strip())

            # ищем до матрацев
            if name[:6].lower() == "матрац":
                break

            # подставляем метку серии в название
            type_product = ""
            if c == 4:  # эко
                type_product = " э"
            if c == 5:  # лайт
                type_product = " л"

            i_product = {
                'name': name + type_product
            }

            # проверяем на повторение
            for i in d_pricelist:
                if i['name'] == i_product['name']:
                    break
            else:  # добавляем только уникальные позиции

                price_doc = ws.cell(row=r, column=c).value or 0

                if price_doc:
                    i_product['price_doc'] = price_doc
                    d_pricelist.append(i_product)

    return d_pricelist


def import_pricelist_code():

    log = [f"Импорт прайса - начало"]

    d_pricelist = get_d_pricelist()

    n_count_added = 0
    n_count_changed = 0

    # добавляем данные в базу данных
    from product.models import Product
    from django.contrib.auth.models import User
    me = User.objects.get(username='alex')

    for i_pricelist in d_pricelist:
        name = i_pricelist['name']
        price_doc = i_pricelist['price_doc']
        if not len(Product.objects.filter(name=name)):
            Product.objects.create(author=me, name=name, price_doc=price_doc)
            n_count_added += 1
        else:
            i_product = Product.objects.get(name=name)
            if price_doc != i_product.price_doc:
                if get_setting_rewrite():
                    log.append(f"Изменение: В файле эксель цена = {price_doc},"
                               f" а в базе данных цена = {i_product.price_doc} : {name}")
                    i_product.price_doc = price_doc
                    i_product.save()
                    n_count_changed += 1

    log.append(f"Импорт окончен. Всего изделий в прайсе: {len(d_pricelist)},"
               f" из них импортировано: {n_count_added}, изменено: {n_count_changed}")
    return log


def get_d_product_calculation():

    f = r"static/data/себест общ новая.xlsx"

    d_product = []

    wb = openpyxl.load_workbook(filename=f, read_only=False, data_only=True)

    for sheetname in wb.sheetnames:
        if sheetname[0:7] == "себ эко" or sheetname[0:12] == "себ лайт":
            # Пока только листы эко и лайт
            pass
        else:
            # Остальные листы пока пропускаем
            continue
        ws = wb[sheetname]
        columns = ws.max_column

        for c in range(1, columns + 1):
            name_calc = ws.cell(row=2, column=c).value
            if not name_calc or not isinstance(name_calc, str):
                continue
            name_calc = del_double_space(name_calc.strip())
            if not len(name_calc):
                continue
            # Символ # в начале наименования изделия означает, что изделие не импортируется (например, разрабатывается)
            if name_calc[0] == "#":
                continue

            # подставляем метку серии в название
            product_type = " unknown"
            if sheetname[4:7] == "эко":
                product_type = " э"
            elif sheetname[4:9] == "станд":
                product_type = " стандарт"
            elif sheetname[4:12] == "лайт":
                product_type = " л"
            # elif sheetname[4:7] == "милд":
            #    product_type = " м"
            elif sheetname[4:9] == "устар":
                product_type = " прочее"

            name_calc += product_type

            i_product = {
                'name': name_calc
            }

            # проверяем на повторение
            for i in d_product:
                if i['name'] == i_product['name']:
                    break
            else:  # добавляем только уникальные позиции
                d_product.append(i_product)

    return d_product


def import_product_calculation_code():

    log = [f"Импорт списка изделий из калькуляции - начало"]

    d_product = get_d_product_calculation()

    n_count_added = 0

    # добавляем данные в базу данных
    from product.models import Product
    from django.contrib.auth.models import User
    me = User.objects.get(username='alex')

    for i_product in d_product:
        name = i_product['name']
        if not len(Product.objects.filter(name=name)):
            Product.objects.create(author=me, name=name)
            n_count_added += 1

    log.append(f"Импорт окончен. Всего изделий в калькуляции: {len(d_product)}, из них импортировано: {n_count_added}")
    return log


def test_import_pricelist_code():
    """
        тестирование идентичности данных в прайслисте (файл эксель) и в базе данных
        проверяются следующие ошибки:
        1) в файле эксель есть изделие (name, price_doc), а в базе данных нет
        2) значение price_doc различается
        3) в базе данных есть изделие (name, price_doc), а в файле эксель нет
    """

    from product.models import Product

    log = [f"Тестирование идентичности данных в прайслисте (файл эксель) и в базе данных"]

    d_pricelist = get_d_pricelist()

    n_count = 0

    for i_pricelist in d_pricelist:
        name = i_pricelist['name']
        if not len(Product.objects.filter(name=name)):
            log.append(f"Ошибка: В файле эксель есть изделие, а в базе данных нет: {name}")
            n_count += 1
            continue
        price_doc = i_pricelist['price_doc']
        i_product = Product.objects.get(name=name)
        if price_doc != i_product.price_doc:
            log.append(f"Ошибка: В файле эксель цена = {price_doc}, а в базе данных цена = {i_product.price_doc}:"
                       f" {name}")
            n_count += 1

    products = Product.objects.all()
    for i_product in products:

        # сравниваем только изделия с фактической (задокументированной) ценой
        if i_product.price_doc is None:
            continue

        name = i_product.name
        # ищем простым перебором
        for i_pricelist in d_pricelist:
            if i_pricelist['name'] == name:
                break
        else:
            log.append(f"Ошибка: В базе данных есть изделие, а в файле эксель нет: {name}")
            n_count += 1

    log.append(f"Тестирование окончено. Всего ошибок найдено: {n_count}")
    return log


def test_import_product_calculation_code():
    """
        тестирование идентичности списка изделий в калькуляции (файл эксель) и в базе данных
        проверяются следующие ошибки:
        1) в файле эксель есть изделие (name), а в базе данных нет
        2) в базе данных есть изделие (name), а в файле эксель нет
    """

    from product.models import Product

    log = [f"Тестирование идентичности списка изделий в калькуляции (файл эксель) и в базе данных"]

    d_product = get_d_product_calculation()

    n_count = 0

    for i_product in d_product:
        name = i_product['name']
        if not len(Product.objects.filter(name=name)):
            log.append(f"Ошибка: В файле эксель есть изделие, а в базе данных нет: {name}")
            n_count += 1
            continue

    products = Product.objects.all()
    for i_product_calc in products:

        # сравниваем только изделия с рассчётной ценой
        if i_product_calc.price_calc is None:
            continue

        name = i_product_calc.name

        # ищем простым перебором
        for i_product in d_product:
            if i_product['name'] == name:
                break
        else:
            log.append(f"Ошибка: В базе данных есть изделие, а в файле эксель нет: {name}")
            n_count += 1

    log.append(f"Тестирование окончено. Всего ошибок найдено: {n_count}")
    return log


def get_d_cost(log):

    f = r"static/data/себест общ новая.xlsx"

    d_cost = []

    wb = openpyxl.load_workbook(filename=f, read_only=False, data_only=True)

    for sheetname in wb.sheetnames:
        if sheetname[0:4] != "себ ":
            continue
        ws = wb[sheetname]
        rows = ws.max_row
        for r in range(1, rows + 1):
            name = ws.cell(row=r, column=1).value
            price = ws.cell(row=r, column=2).value or 0
            waste_percent = ws.cell(row=r, column=3).value or 0
            if not name or not isinstance(name, str):
                continue
            name = del_double_space(name.strip())
            if name == "себестоимость":
                break
            i_cost = {
                'name': name,
                'price': price,
                'waste_percent': waste_percent
            }
            # проверяем на повторение
            for i in d_cost:
                if i['name'] == i_cost['name']:
                    if i['price'] != i_cost['price']:
                        log.append(f"Цена у материала {i['name']} отличается {i['price']} != {i_cost['price']}")
                    if i['waste_percent'] != i_cost['waste_percent']:
                        log.append(f"Доля потерь(обрези) у материала {i['name']} отличается "
                                   f"{i['waste_percent']} != {i_cost['waste_percent']}")
                    break
            else:  # добавляем только уникальные позиции
                d_cost.append(i_cost)

    return d_cost


def import_cost_code():

    log = [f"Импорт справочника затрат - начало"]

    n_count_added = 0
    n_count_changed = 0

    d_cost = get_d_cost(log)

    # добавляем данные в базу данных
    from product.models import Cost

    sort = 0
    for i_cost in d_cost:
        sort += 10
        name = i_cost['name']
        price = i_cost['price']
        waste_percent = i_cost['waste_percent']
        if not len(Cost.objects.filter(name=name)):
            Cost.objects.create(name=name, price=price, waste_percent=waste_percent, sort=sort)
            n_count_added += 1
        else:
            i_cost_edit = Cost.objects.get(name=name)
            if price != i_cost_edit.price:
                if get_setting_rewrite():
                    log.append(f"Ошибка: В файле эксель цена = {price}, а в базе данных цена = {i_cost_edit.price}:"
                               f" {name}")
                    i_cost_edit.price = price
                    i_cost_edit.save()
                    n_count_changed += 1

    log.append(f"Импорт справочника затрат окончен. Всего затрат в калькуляции: {len(d_cost)}, "
               f"из них импортировано: {n_count_added}, изменено: {n_count_changed}")
    return log


def test_import_cost_code():
    """
        тестирование идентичности данных в калькуляции (файл эксель) и в базе данных
        проверяются следующие ошибки:
        1) в файле эксель есть статья затрат, а в базе данных нет
        2) значение price различается
        3) в базе данных есть статья затрат, а в файле эксель нет
    """

    from product.models import Cost

    log = [f"Тестирование идентичности статей затрат в калькуляции (файл эксель) и в базе данных"]

    d_cost_calc = get_d_cost(log)

    n_count = 0

    for i_cost_calc in d_cost_calc:
        name = i_cost_calc['name']
        if not len(Cost.objects.filter(name=name)):
            log.append(f"Ошибка: В файле эксель есть статья затрат, а в базе данных нет: {name}")
            n_count += 1
            continue
        price = i_cost_calc['price']
        waste_percent = i_cost_calc['waste_percent']
        i_cost = Cost.objects.get(name=name)
        if price != i_cost.price:
            log.append(f"Ошибка: В файле эксель цена = {price}, а в базе данных цена = {i_cost.price}:"
                       f" {name}")
            n_count += 1
        if waste_percent != i_cost.waste_percent:
            log.append(f"Ошибка: В файле эксель процент отходов = {waste_percent}, "
                       f"а в базе данных цена = {i_cost.waste_percent}: {name}")
            n_count += 1

    products = Cost.objects.all()
    for i_product in products:
        name = i_product.name
        # ищем простым перебором
        for i_cost_calc in d_cost_calc:
            if i_cost_calc['name'] == name:
                break
        else:
            log.append(f"Ошибка: В базе данных есть статья затрат, а в файле эксель нет: {name}")
            n_count += 1

    log.append(f"Тестирование окончено. Всего ошибок найдено: {n_count}")
    return log


def get_d_calculation(log, add_price_calc=False):

    # if add_price_calc = True - добавляем еще информацию о расчетных ценах в файле

    from product.models import Product, Cost

    f = r"static/data/себест общ новая.xlsx"

    d_calculation = []
    d_product_price_calc = []  # расчетная цена в файле эксель

    wb = openpyxl.load_workbook(filename=f, read_only=False, data_only=True)

    for sheetname in wb.sheetnames:

        if sheetname[0:7] == "себ эко" or sheetname[0:12] == "себ лайт":
            # Пока только листы эко и лайт
            pass
        else:
            # Остальные листы пока пропускаем
            continue

        ws = wb[sheetname]
        columns = ws.max_column
        rows = ws.max_row

        for c in range(1, columns + 1):
            name_calc = ws.cell(row=2, column=c).value
            if not name_calc or not isinstance(name_calc, str):
                continue
            name_calc = del_double_space(name_calc.strip())
            if not len(name_calc):
                continue
            # Символ # в начале наименования изделия означает, что изделие не импортируется (например, разрабатывается)
            if name_calc[0] == "#":
                continue

            # подставляем метку серии в название
            product_type = " unknown"
            if sheetname[4:7] == "эко":
                product_type = " э"
            elif sheetname[4:9] == "станд":
                product_type = " стандарт"
            elif sheetname[4:12] == "лайт":
                product_type = " л"
            elif sheetname[4:7] == "мдф":
                product_type = " мдф"
            elif sheetname[4:9] == "устар":
                product_type = " прочее"

            name_calc += product_type

            # код изделия
            try:
                product_id = Product.objects.get(name=name_calc)
            except Product.DoesNotExist:
                log.append(f"Ошибка: В файле эксель изделие, которого нет в справочнике изделий: {name_calc}")
                continue

            # проверяем на повторение
            for i in d_calculation:
                if i['product_id'] == product_id:
                    log.append(f"Ошибка: В файле эксель повторяется изделие: {name_calc}")
                    break
            else:  # добавляем только уникальные позиции

                for r in range(1, rows + 1):
                    name = ws.cell(row=r, column=1).value

                    if not name or not isinstance(name, str):
                        continue
                    name = del_double_space(name.strip())

                    # условие выхода их цикла - статьи кончились
                    if name == "себестоимость":

                        # если нужно выводить дополнительно данные о рассчитанных ценах в файле эксель
                        if add_price_calc:
                            i_product_price_calc = {
                                'product_id': product_id,
                                'price_calc': ws.cell(row=r, column=c+1).value or 0
                            }
                            d_product_price_calc.append(i_product_price_calc)

                        break

                    # код статьи расходов
                    try:
                        cost_id = Cost.objects.get(name=name)
                    except Cost.DoesNotExist:
                        log.append(f"Ошибка: В файле эксель статья затрат, которой нет в справочнике: {name}")
                        continue

                    amount = ws.cell(row=r, column=c).value or 0
                    cost_add = 0
                    if not amount:
                        cost_add = ws.cell(row=r, column=c+1).value or 0

                    if amount or cost_add:

                        i_calculation = {
                            'product_id': product_id,
                            'cost_id': cost_id,
                            'amount': amount,
                            'cost_add': cost_add
                        }

                        d_calculation.append(i_calculation)

    if add_price_calc:
        return d_calculation, d_product_price_calc
    else:
        return d_calculation


def import_calculation_code():
    log = [f"Импорт калькуляции изделий - начало"]

    n_count_added = 0

    d_calculation = get_d_calculation(log)

    # добавляем данные в базу данных
    from product.models import Calculation

    for i_calculation in d_calculation:
        product_id = i_calculation['product_id']
        cost_id = i_calculation['cost_id']
        amount = i_calculation['amount']
        cost_add = i_calculation['cost_add']
        if not len(Calculation.objects.filter(product_id=product_id, cost_id=cost_id)):
            Calculation.objects.create(product_id=product_id, cost_id=cost_id, amount=amount, cost_add=cost_add)
            n_count_added += 1

    log.append(f"Импорт калькуляции изделий окончен. Всего строк в калькуляции: {len(d_calculation)}, "
               f"из них импортировано: {n_count_added}")
    return log


def test_import_calculation_code():
    """
        тестирование идентичности данных в калькуляции (файл эксель) и в базе данных
        проверяются следующие ошибки:
        1) в файле эксель есть статья затрат, а в базе данных нет
        2) значение amount различается
        3) значение cost_add различается
        4) в базе данных есть статья затрат, а в файле эксель нет
    """
    from product.models import Calculation, Product

    log = [f"Тестирование идентичности строк калькуляции (файл эксель) и в базе данных"]

    d_calculation, d_product_price_calc = get_d_calculation(log, add_price_calc=True)

    # начальное количество найденных ошибок равно количеству строк об ошибках, добавленных в вызываемой ранее функции
    # минус строка-заголовок (1 штука)
    n_count = len(log) - 1

    for i_calculation in d_calculation:
        product_id = i_calculation['product_id']
        cost_id = i_calculation['cost_id']
        amount = i_calculation['amount']
        cost_add = i_calculation['cost_add']
        if not len(Calculation.objects.filter(product_id=product_id, cost_id=cost_id)):
            log.append(f"Ошибка: В файле эксель есть статья затрат для {product_id.name},"
                       f" а в базе данных нет: {cost_id.name}")
            n_count += 1
            continue

        i_calculation_db = Calculation.objects.get(product_id=product_id, cost_id=cost_id)
        if amount != i_calculation_db.amount:
            log.append(f"Ошибка: В файле эксель количество = {amount}, "
                       f"а в базе данных количество = {i_calculation_db.amount} "
                       f"для {product_id.name} статья: {cost_id.name} (calculation.id = {i_calculation_db.id})")
            n_count += 1
        if cost_add != i_calculation_db.cost_add:
            log.append(f"Ошибка: В файле эксель расходы в денежном выражении = {cost_add}, "
                       f"а в базе данных расходы = {i_calculation_db.cost_add} "
                       f"для {product_id.name} статья: {cost_id.name} (calculation.id = {i_calculation_db.id})")
            n_count += 1

    d_calculation_db = Calculation.objects.all()
    for i_calculation_db in d_calculation_db:
        product_id = i_calculation_db.product_id
        cost_id = i_calculation_db.cost_id
        # ищем простым перебором
        for i_calculation in d_calculation:
            if i_calculation['product_id'] == product_id and i_calculation['cost_id'] == cost_id:
                break
        else:
            log.append(f"Ошибка: В базе данных есть статья затрат для {product_id.name},"
                       f" а в файле эксель нет: {cost_id.name}")
            n_count += 1

    # проверка рассчитанных цен в файле эксель и в базе данных
    for i_product_price_calc in d_product_price_calc:
        i_product_db = Product.objects.get(id=i_product_price_calc['product_id'].id)
        if i_product_db is not None:
            price_calc_db = round(float(i_product_db.price_calc), 2)
        else:
            price_calc_db = 0
        # TODO: вынести коэффициент в настройки, пока в целях отладки прописываем в коде
        price_calc_file = round(i_product_price_calc['price_calc'] * 1.43 * 1.4, 2)

        if price_calc_db != price_calc_file:
            log.append(f"Ошибка: В файле эксель у изделия {i_product_db.name} "
                       f"рассчитанная цена = {price_calc_file} а в базе данных = {price_calc_db}")

    log.append(f"Тестирование окончено. Всего ошибок найдено: {n_count}")
    return log


def re_price_calc_code(mode="save"):

    from product.models import Calculation, Product

    if mode == "test":
        log = [f"Тестирование расчетных цен - начало"]
    else:
        log = [f"Пересчет расчетных цен - начало"]

    n_count = 0

    calculations = Calculation.objects.all()

    items = []

    for calculation in calculations:

        # Смотрим указана ли для данного изделия уточненная доля отходов
        if calculation.waste_percent:  # да, указана
            waste_percent_fact = calculation.waste_percent
        else:  # нет, не указано
            waste_percent_fact = calculation.cost_id.waste_percent

        summ = (calculation.cost_id.price or 0) * (1 + (waste_percent_fact or 0)) * (calculation.amount or 0)
        summ += calculation.cost_add or 0

        item = {
            'product_id': calculation.product_id,
            'price_calc': summ
        }

        # ищем простым перебором
        for i_item in items:
            if i_item['product_id'] == item['product_id']:
                i_item['price_calc'] += item['price_calc']
                break
        else:
            items.append(item)

    # коэффициент наценки для получения оптовой цены - цена со склада
    ratio_storage = 1.43
    # коэффициент наценки для получения розничной цены - цена в магазине
    ratio_shop = 1.4
    # общий коэффициент
    ratio = ratio_storage * ratio_shop

    for i_item in items:
        i_product = Product.objects.get(pk=i_item['product_id'].id)
        new_price = round(i_item['price_calc'] * ratio, 2)
        old_price = round(float(i_product.price_calc or 0), 2)
        if mode == "test":
            if new_price != old_price:
                log.append(f"Ошибка:  у изделия {i_product.name}  рассчитанная цена = {new_price}"
                           f" а в базе данных сейчас = {old_price}")
        else:
            i_product.price_calc = new_price
            i_product.save()
        n_count += 1

    if mode == "test":
        log.append(f"Тестирование расчетных цен окончено. Всего цен проверено: {n_count}")
    else:
        log.append(f"Пересчет расчетных цен окончен. Всего цен пересчитано: {n_count}")

    return log
